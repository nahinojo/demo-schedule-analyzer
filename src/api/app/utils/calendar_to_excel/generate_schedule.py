from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

from app import PATH_TO_SCHEDULE
from .date_difference_school_weeks import date_difference_school_weeks
from .extract_courses_from_calendar import extract_courses_from_calendar


def generate_schedule(
        target_course_code: str = None,
        target_instructor: str = None,
        target_term: str = None,
        target_year: int = 2023,
        is_target_year_as_minium: bool = True,
):
    """
    Generates a schedule.

    Schedule is stored in an .xlsx file at `SCHEDULE_PATH`.

    Parameters
    ----------
    target_course_code: str
        The target course code.
    target_instructor: str
        The target instructor.
    target_term: str
        The target term.
    target_year: int
        The target year.
    is_target_year_as_minium: bool
        Whether the target year is the minimum year.

    Returns
    -------
    None
    """
    # !!!Should be pulling from the database, not the function.
    courses = extract_courses_from_calendar(
        target_course_code=target_course_code,
        target_instructor=target_instructor,
        target_term=target_term,
        target_year=target_year,
    )
    wb = Workbook()
    ws = wb.active  # type: Worksheet
    demo_count_list = []
    for course_idx, course in enumerate(courses):
        ws.title = (
            f"{course.course_code}"
            f" {course.instructor}"
            f" - {course.term}"
            f" {course.year}"
        )
        ws["A1"] = "Week - Day"
        ws["B1"] = "Demonstrations"
        ws["C1"] = "Additional Information"
        ws["D1"] = "Course Information"
        ws.merge_cells("D1:E1")
        ws["D2"] = "Instructor"
        ws["E2"] = course.instructor
        ws["D3"] = "Course Code"
        ws["E3"] = course.course_code
        ws["D4"] = "Term"
        ws["E4"] = course.term
        ws["D5"] = "Year"
        ws["E5"] = course.year

        course_demos_list = []
        row_idx = 2
        first_demo_event_date = course.demo_events[0].event_date
        demo_event_week = int(
            first_demo_event_date.month != 9
            or first_demo_event_date.weekday() < 3
        )  # Fall quarter begins at Week 0
        demo_count = 0
        for demo_event_idx, demo_event in enumerate(course.demo_events):
            if demo_event_idx > 0:
                demo_event_week += date_difference_school_weeks(
                    prev_date=course.demo_events[demo_event_idx - 1].event_date,
                    next_date=demo_event.event_date
                )
            demo_event_date = demo_event.event_date
            demo_event_weekday = demo_event_date.strftime("%A")
            ws[f"A{row_idx}"] = f"{demo_event_week} - {demo_event_weekday}"
            ws[f"C{row_idx}"] = demo_event.additional_information
            init_row_idx = row_idx
            for demo_idx, demo in enumerate(demo_event.demos):
                ws[f"B{row_idx}"] = demo.name
                row_idx += 1
                demo_count += 1
            demo_event_span_rows = row_idx - init_row_idx
            if demo_event_span_rows > 1:
                ws.merge_cells(f"A{init_row_idx}:A{row_idx - 1}")
                ws.merge_cells(f"C{init_row_idx}:C{row_idx - 1}")
        demo_count_list.append(demo_count)
        if course_idx < len(courses) - 1:
            wb.create_sheet("next")
            ws = wb["next"]

    default_font_name = "Ubuntu"
    default_font_size = 12
    default_indent = 1
    default_vertical = "center"
    default_horizontal = "left"
    default_fill_type = "solid"
    dark_blue = "0D47A1"
    light_blue = "DAE3F3"
    dark_green = "004D40"
    light_green = "DBE6D5"
    default_side = Side(
        style="thin",
        color="FFFFFF"
    )
    top_border = Border(
        top=default_side
    )
    bottom_border = Border(
        bottom=default_side
    )
    outline_border = Border(
        top=default_side,
        bottom=default_side,
        left=default_side,
        right=default_side,
    )
    default_font = Font(
        name=default_font_name,
        size=default_font_size
    )
    bold_font = Font(
        bold=True,
        name=default_font_name,
        size=default_font_size
    )
    default_alignment = Alignment(
        vertical=default_vertical,
        horizontal=default_horizontal,
        indent=default_indent,
        relativeIndent=default_indent,
        shrink_to_fit=False
    )
    wrap_alignment = Alignment(
        vertical=default_vertical,
        horizontal=default_horizontal,
        indent=default_indent,
        wrap_text=True,
        shrink_to_fit=False
    )
    absolute_center_alignment = Alignment(
        vertical=default_vertical,
        horizontal="center",
        shrink_to_fit=False
    )
    title_font = Font(
        bold=True,
        color="FFFFFF",
        name=default_font_name,
        size=14,
    )
    demo_event_title_fill = PatternFill(
        fgColor=dark_blue,
        fill_type=default_fill_type
    )
    demo_event_data_fill = PatternFill(
        fgColor=light_blue,
        fill_type=default_fill_type,
    )
    course_information_title_fill = PatternFill(
        fgColor=dark_green,
        fill_type=default_fill_type
    )
    course_information_data_fill = PatternFill(
        fgColor=light_green,
        fill_type=default_fill_type
    )

    for sheet_idx, sheet in enumerate(wb.worksheets):
        sheet.row_dimensions[1].height = 45
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 45
        sheet.column_dimensions["C"].width = 30
        sheet.column_dimensions["D"].width = 15
        sheet.column_dimensions["E"].width = 15
        is_new_date = True
        max_row = demo_count_list[sheet_idx] + 1
        if max_row < 5:
            max_row = 5
        for row_idx_0, row in enumerate(sheet.iter_rows(
            min_col=1,
            max_col=5,
            max_row=max_row
        )):
            row_idx = row_idx_0 + 1  # 1-index conversion for Sheet.
            if row_idx > 1:
                sheet.row_dimensions[row_idx].height = 40
            weekday_cell = row[0]
            is_empty_weekday_cell = weekday_cell.value is None
            is_new_date = not is_empty_weekday_cell
            is_leave_demo_events_empty = is_empty_weekday_cell and type(weekday_cell) != MergedCell
            for col_idx_0, cell in enumerate(row):
                col_idx = col_idx_0 + 1  # 1-index conversion for Sheet.
                is_title = row_idx == 1
                is_demo_events = col_idx <= 3
                is_weekday = col_idx == 1
                is_demos = col_idx == 2
                is_additonal_information = col_idx == 3
                is_course_information = col_idx >= 4 and row_idx <= 5
                is_course_information_label = col_idx == 4
                is_course_information_data = col_idx == 5
                if is_title:
                    cell.alignment = absolute_center_alignment
                    cell.font = title_font
                    if is_course_information:
                        cell.fill = course_information_title_fill
                    else:
                        cell.fill = demo_event_title_fill
                else:
                    if is_demo_events and not is_leave_demo_events_empty:
                        cell.fill = demo_event_data_fill
                        if is_weekday:
                            cell.alignment = absolute_center_alignment
                            cell.font = bold_font
                            if is_new_date:
                                cell.border = outline_border
                        elif is_demos:
                            cell.alignment = default_alignment
                            cell.font = default_font
                            if is_new_date:
                                cell.border = top_border
                        elif is_additonal_information:
                            cell.alignment = wrap_alignment
                            cell.font = default_font
                            if is_new_date:
                                cell.border = outline_border
                    elif is_course_information:
                        cell.fill = course_information_data_fill
                        cell.border = outline_border
                        if is_course_information_label:
                            cell.alignment = default_alignment
                            cell.font = bold_font
                        else:
                            cell.alignment = default_alignment
                            cell.font = default_font
    wb.save(PATH_TO_SCHEDULE)
