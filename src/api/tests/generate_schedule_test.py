import pytest


@pytest.mark.usefixtures("app_context_with_real_data")
def generate_schedule_test():
    """
    Tests generate_schedule function.
    """
    from app import PATH_TO_SCHEDULE
    from app.database import Session
    from app.models import Course
    from app.utils import generate_schedule
    from openpyxl import load_workbook, Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell
    from openpyxl.styles.colors import Color
    from sqlalchemy import select

    with Session() as session:
        num_courses = len(session.execute(select(Course)).scalars().all())
        assert num_courses > 10
        assert num_courses < 1000
    for i in range(1, num_courses + 1):
        generate_schedule([i])
        wb: Workbook = load_workbook(PATH_TO_SCHEDULE)
        assert len(wb.sheetnames) == 1
        ws: Worksheet = wb[wb.sheetnames[0]]
        for col in ["A", "B", "C", "D", "E"]:
            row_idx = 1
            cell: Cell = ws[f"{col}{row_idx}"]
            cell_fg_color: Color = cell.fill.fgColor.value
            if cell.value:
                assert cell_fg_color != '00000000'
    assert True
