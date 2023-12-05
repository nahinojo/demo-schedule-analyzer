from utils import (
    date_difference_school_weeks,
)
from openpyxl import Workbook
from parse_calendar import parse_calendar

BUILD_PATH = "./build"
SCHEDULE_FILE_NAME = "demo-schedule.xlsx"
SCHEDULE_PATH = f"{BUILD_PATH}/{SCHEDULE_FILE_NAME}"


def generate_schedule(course_details_list: list):
    wb = Workbook()
    ws = wb.active
    for idx, course_details in course_details_list:
        ws.title = (
            f"{course_details['course_code']}"
            f" {course_details['instructor']}"
            f" - {course_details['term']}"
            f" {course_details['year']}"
        )
        cell_a1 = ws["A1"]
        cell_a1.value = "Week - Day"
        ws["B1"] = "Demos"
        ws["C1"] = "Additional Information"
        ws.merge_cells("D1:E1")
        ws["D1"] = "Course Information"
        number_of_demos_in_course = 0
        for demo_event in course_details["demo_event_list"]:
            number_of_demos_in_course += len(demo_event["demos"])
        for cell_demo in ws.iter_rows(
            min_col=2,
            max_col=2,
            max_row=number_of_demos_in_course
        ):
            pass

        earliest_demo_event_date = course_details["demo_event_list"][0]["date"]
        # latest_demo_event_date = course_details["demo_event_list"][-1]["date"]
        # schedule_time_span_school_weeks = (
        #         1 + date_difference_school_weeks(
        #             latest_demo_event_date,
        #             earliest_demo_event_date
        #         )
        # )
        school_week_start = 1
        if course_details["term"] == "Fall" and earliest_demo_event_date.month == 9:
            school_week_start = 0

        if idx < len(course_details_list) - 1:
            wb.create_sheet("next")
            ws = wb["next"]

    wb.save(SCHEDULE_PATH)


if __name__ == "__main__":
    test_course_details_list = parse_calendar(
        target_year=2023,
        target_instructor="Ochoa"
    )
    generate_schedule(test_course_details_list)
