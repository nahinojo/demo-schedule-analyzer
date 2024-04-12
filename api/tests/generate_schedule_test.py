import pytest


@pytest.mark.usefixtures("app_context_with_real_data")
def generate_schedule_test():
    """
    Tests generate_schedule function. The following cases are tested:
    - A schedule sheet can be created for each Course.
    - Every cell with a value is colored
    - Every cell without a value is not colored.
    """
    from typing import cast

    from app import PATH_TO_SCHEDULE
    from app.database import Session
    from app.models import Course
    from app.utils import generate_schedule
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.styles.colors import Color
    from openpyxl.worksheet.worksheet import Worksheet
    from sqlalchemy import select

    with Session() as session:
        num_courses = len(session.execute(select(Course)).scalars().all())
    for i in range(1, num_courses + 1):
        generate_schedule([i])
        wb = cast(Workbook, load_workbook(PATH_TO_SCHEDULE))
        assert len(wb.sheetnames) == 1
        ws = cast(Worksheet, wb[wb.sheetnames[0]])
        for col in ["A", "B", "C", "D", "E"]:
            row_idx = 1
            cell = cast(Cell, ws[f"{col}{row_idx}"])
            cell_value = cell.value
            cell_fg_color = cast(Color, cell.fill.fgColor.value)
            is_cell_colored = cell_fg_color != "00000000"
            is_value_and_color = cell_value and is_cell_colored
            is_not_value_and_color = not cell_value and not is_cell_colored
            is_value_xnor_color = is_value_and_color or is_not_value_and_color
            assert is_value_xnor_color
    assert True
